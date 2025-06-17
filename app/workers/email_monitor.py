from celery import Celery
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from app.database import get_db, engine
from app.models import User, EmailProvider, Email, Attachment
from app.email import GmailProvider, OutlookProvider
from app.triggers.handler import TriggerHandler
from app.workers.celery_app import celery_app
from app.document_ai.client import DocumentAIClient
import redis
import json
from typing import List, Dict, Any
from app.providers import EmailProviderFactory
from openpyxl import Workbook
import os
from pathlib import Path
from openpyxl import load_workbook

# Initialize Document AI client
document_ai_client = DocumentAIClient()

redis_client = redis.from_url("redis://redis:6379/0")

# Placeholder for Document AI processor IDs for extraction
# You will need to replace these with your actual processor IDs for each document type
DOCUMENT_AI_EXTRACTION_PROCESSOR_IDS = {
    "invoice_turkey": os.getenv("DOCUMENT_AI_INVOICE_PROCESSOR_ID", "6ffa074f060f9676"),
    "export_declaration_turkey_house": os.getenv("DOCUMENT_AI_EXPORT_PROCESSOR_ID", "1b500327b4b8d7a2"),
    "invoice_generic": os.getenv("DOCUMENT_AI_COMMERCIAL_INVOICE_PROCESSOR_ID", "4b86b4ff97c89de0"),
    "consignment_instructions": os.getenv("DOCUMENT_AI_CONSIGNMENT_INSTRUCTIONS_PROCESSOR_ID", "d8a2e50820b7b484")    # Add more document types and their corresponding processor IDs here
}

@celery_app.task
def check_all_users_emails():
    """Background task to check all users' emails for new messages"""
    db = Session(bind=engine)
    try:
        active_providers = db.query(EmailProvider).filter(
            EmailProvider.is_active == True
        ).all()
        
        for provider in active_providers:
            check_user_emails.delay(provider.id)
    
    finally:
        db.close()


@celery_app.task
def check_user_emails(provider_id: int):
    """Check emails for a specific user provider"""
    db = Session(bind=engine)
    try:
        provider = db.query(EmailProvider).filter(
            EmailProvider.id == provider_id
        ).first()
        
        if not provider or not provider.is_active:
            return
        
        # Get the appropriate email provider
        email_provider = get_email_provider_instance(provider)
        if not email_provider:
            return
        
        # Get last sync time or default to 1 hour ago
        last_sync = provider.last_sync or (datetime.now() - timedelta(hours=1))
        
        # Fetch new emails since last sync
        new_messages = email_provider.get_new_messages_since(last_sync)
        
        if new_messages:
            print(f"Found {len(new_messages)} new messages for user {provider.user_id}")
            
            # Process each new message
            for message in new_messages:
                process_new_email.delay(provider.user_id, provider.id, message.__dict__)
        
        # Update last sync time
        provider.last_sync = datetime.now()
        db.commit()
    
    except Exception as e:
        print(f"Error checking emails for provider {provider_id}: {e}")
        db.rollback()
    
    finally:
        db.close()


@celery_app.task
def process_new_email(user_id: int, provider_id: int, message_data: dict):
    """Process a new email message and trigger handlers"""
    try:
        print(f"📧 Processing new email for user {user_id}: {message_data.get('subject')}")
        
        # Debug log the incoming message data
        print(f"📦 Raw message data: {json.dumps(message_data, default=str)}")
        
        # Convert datetime objects to ISO format strings for JSON serialization
        if 'received_at' in message_data and isinstance(message_data['received_at'], datetime):
            message_data['received_at'] = message_data['received_at'].isoformat()
            print(f"📅 Converted received_at to ISO format: {message_data['received_at']}")
        
        # Store the message in Redis for quick access
        try:
            message_key = f"email:user:{user_id}:message:{message_data['id']}"
            redis_client.setex(message_key, 3600, json.dumps(message_data))  # Store for 1 hour
            print(f"💾 Stored email in Redis with key: {message_key}")
        except Exception as e:
            print(f"❌ Failed to store in Redis: {e}")
        
        # Get trigger handler and process the email
        try:
            trigger_handler = TriggerHandler()
            trigger_handler.handle_new_email(user_id, provider_id, message_data)
            print(f"✅ Trigger processing completed")
        except Exception as e:
            print(f"❌ Trigger processing failed: {e}")
        
        # Save email to database
        db = Session(bind=engine)
        try:
            print(f"🔍 Checking for existing email in database...")
            # Check if email already exists to avoid duplicates
            existing_email = db.query(Email).filter(
                Email.external_id == message_data['id'],
                Email.user_id == user_id
            ).first()
            
            if not existing_email:
                print(f"📝 Email not found in database, creating new record...")
                # Get provider info
                provider_db = db.query(EmailProvider).filter(
                    EmailProvider.id == provider_id
                ).first()
                
                if provider_db:
                    # Get the actual provider instance for fetching attachments
                    provider = EmailProviderFactory.create_email_provider(
                        provider_db.provider_type,
                        provider_db.access_token,
                        provider_db.refresh_token
                    )
                    
                    # Convert ISO string back to datetime for database storage
                    received_at = None
                    if message_data.get('received_at'):
                        received_at = datetime.fromisoformat(message_data['received_at'].replace('Z', '+00:00'))
                        print(f"📅 Converted received_at back to datetime: {received_at}")
                    
                    # Ensure recipients and labels are lists
                    recipients = message_data.get('recipients', [])
                    if isinstance(recipients, str):
                        recipients = [recipients]
                    elif recipients is None:
                        recipients = []
                    print(f"📨 Recipients: {recipients}")
                    
                    labels = message_data.get('labels', [])
                    if isinstance(labels, str):
                        labels = [labels]
                    elif labels is None:
                        labels = []
                    print(f"🏷️ Labels: {labels}")
                    
                    # Create the email record first
                    new_email = Email(
                        external_id=message_data['id'],
                        user_id=user_id,
                        provider_id=provider_id,
                        subject=message_data.get('subject'),
                        sender=message_data.get('sender'),
                        recipients=recipients,
                        thread_id=message_data.get('thread_id'),
                        body=message_data.get('body', ''),
                        html_body=message_data.get('html_body'),
                        is_read=message_data.get('is_read', False),
                        labels=labels,
                        received_at=received_at
                    )
                    
                    # Save the email first to get its ID
                    db.add(new_email)
                    db.flush()  # This will generate the ID without committing
                    
                    # Process attachments if any
                    if message_data.get('attachments'):
                        print(f"📎 Processing {len(message_data['attachments'])} attachments")
                        print(f"📎 Attachment data: {json.dumps(message_data['attachments'], default=str)}")
                        for attachment in message_data['attachments']:
                            print(f"📎 Processing attachment: {attachment['filename']}")
                            attachment_data = provider.get_attachment_data(message_data['id'], attachment['id'])
                            if attachment_data:
                                print(f"📎 Got attachment data, size: {len(attachment_data)} bytes")
                                
                                # Classify the document if it's a supported type
                                classification = None
                                if attachment['content_type'] in ['application/pdf', 'image/jpeg', 'image/png', 'image/tiff']:
                                    try:
                                        print(f"🔍 Classifying document: {attachment['filename']}")
                                        classification = document_ai_client.classify_document(
                                            attachment_data,
                                            attachment['content_type']
                                        )
                                        print(f"✅ Document classification result: {json.dumps(classification, default=str)}")
                                    except Exception as e:
                                        print(f"❌ Document classification failed: {str(e)}")
                                        classification = {
                                            "error": str(e),
                                            "confidence": 0.0,
                                            "type": "error"
                                        }
                                
                                new_attachment = Attachment(
                                    email_id=new_email.id,
                                    external_id=attachment['id'],
                                    filename=attachment['filename'],
                                    content_type=attachment['content_type'],
                                    size=attachment['size'],
                                    data=attachment_data,
                                    is_downloaded=True,
                                    # Add classification results if available
                                    document_type=classification.get('type') if classification else None,
                                    classification_confidence=classification.get('confidence') if classification else None,
                                    page_count=classification.get('page_count') if classification else None,
                                    classification_error=classification.get('error') if classification else None,
                                    classification_metadata=classification if classification else None
                                )
                                new_email.attachments.append(new_attachment)
                                print(f"📎 Added attachment to email: {attachment['filename']}")
                            else:
                                print(f"❌ Failed to get attachment data for: {attachment['filename']}")
                    else:
                        print("ℹ️ No attachments found in message")
                    
                    # Now commit everything together
                    db.commit()
                    print(f"💾 Email saved to database: {message_data.get('subject')}")
                else:
                    print(f"❌ Provider not found for user {user_id}")
            else:
                print(f"⚠️ Email already exists in database: {message_data.get('subject')}")
        
        except Exception as e:
            print(f"❌ Failed to save email to database: {e}")
            import traceback
            traceback.print_exc()
            db.rollback()
        
        finally:
            db.close()
        
        print(f"✅ Processed new email: {message_data.get('subject', 'No Subject')} for user {user_id}")
    
    except Exception as e:
        print(f"❌ Error processing email for user {user_id}: {e}")
        import traceback
        traceback.print_exc()


def get_email_provider_instance(provider: EmailProvider):
    """Get the appropriate email provider instance"""
    try:
        if provider.provider_type == "gmail":
            return GmailProvider(
                access_token=provider.access_token,
                refresh_token=provider.refresh_token
            )
        elif provider.provider_type == "outlook":
            return OutlookProvider(access_token=provider.access_token)
        else:
            print(f"Unsupported provider type: {provider.provider_type}")
            return None
    
    except Exception as e:
        print(f"Error creating email provider instance: {e}")
        return None


@celery_app.task
def process_attachments():
    """Task to extract data from classified attachments using Document AI and save to Excel.
    Continuously writes to current week's Excel file and creates new one when week changes.
    """
    db = Session(bind=engine)
    output_dir = Path("output/excel_reports")
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        # Define cutoff for attachments not extracted in the last week
        one_week_ago = datetime.now() - timedelta(weeks=1)

        # Query attachments that have a document_type, have data, and haven't been extracted recently
        attachments_to_extract = db.query(Attachment).filter(
            Attachment.document_type.isnot(None),
            Attachment.data.isnot(None),
            (Attachment.last_extracted_at.is_(None) | (Attachment.last_extracted_at < one_week_ago))
        ).all()

        if not attachments_to_extract:
            print("ℹ️ No new classified attachments found for extraction.")
            return

        print(f"🔍 Found {len(attachments_to_extract)} classified attachments for extraction.")

        # Print available document types and their processor IDs
        print("\nAvailable document types and their processor IDs:")
        for doc_type, processor_id in DOCUMENT_AI_EXTRACTION_PROCESSOR_IDS.items():
            print(f"  - {doc_type}: {processor_id}")

        # Group attachments by document_type to use correct processors
        attachments_by_type = {}
        unsupported_attachments = []
        
        for att in attachments_to_extract:
            print(f"\nProcessing attachment: {att.filename}")
            print(f"Document type from classifier: {att.document_type}")
            print(f"Classification confidence: {att.classification_confidence}")
            
            if att.document_type and att.document_type in DOCUMENT_AI_EXTRACTION_PROCESSOR_IDS:
                if att.document_type not in attachments_by_type:
                    attachments_by_type[att.document_type] = []
                attachments_by_type[att.document_type].append(att)
            else:
                print(f"⚠️ Skipping attachment {att.filename}: No extraction processor defined for type '{att.document_type}'.")
                unsupported_attachments.append(att)
        
        # Mark unsupported attachments as processed to avoid repeated attempts
        if unsupported_attachments:
            print(f"\n📝 Marking {len(unsupported_attachments)} unsupported attachments as processed to avoid repeated attempts:")
            for att in unsupported_attachments:
                att.last_extracted_at = datetime.now()
                print(f"  - {att.filename} ({att.document_type})")

        # Dictionary to store all extracted data, organized by thread_id
        all_thread_data = {}

        for doc_type, attachments in attachments_by_type.items():
            processor_id = DOCUMENT_AI_EXTRACTION_PROCESSOR_IDS[doc_type]
            print(f"Processing {len(attachments)} attachments for document type '{doc_type}' with processor '{processor_id}'...")

            for attachment in attachments:
                try:
                    extraction_result = document_ai_client.extract_document_entities(
                        attachment.data,
                        attachment.content_type,
                        processor_id
                    )

                    if "extracted_data" in extraction_result and extraction_result["extracted_data"]:
                        # Get thread_id and email metadata
                        thread_id = attachment.email.thread_id if attachment.email else "unknown"
                        email_subject = attachment.email.subject if attachment.email else "N/A"
                        email_sender = attachment.email.sender if attachment.email else "N/A"
                        
                        # Initialize thread data if not exists
                        if thread_id not in all_thread_data:
                            all_thread_data[thread_id] = {
                                "Thread ID": thread_id,
                                "Email Subject": email_subject,
                                "Email Sender": email_sender,
                                "Extraction Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "Entities": {}  # Store all entities here
                            }
                        
                        # Add all entities to the thread data with document type prefix
                        for entity_name, entity_value in extraction_result["extracted_data"].items():
                            if entity_value:  # Only add non-empty values
                                # Prefix entity name with document type to differentiate same entities from different doc types
                                prefixed_entity_name = f"{doc_type}_{entity_name}"
                                all_thread_data[thread_id]["Entities"][prefixed_entity_name] = entity_value
                        
                        print(f"✅ Extracted data from {attachment.filename} ({doc_type}) in thread {thread_id}.")
                    else:
                        print(f"❌ No data extracted from {attachment.filename} ({doc_type}). Error: {extraction_result.get('error', 'Unknown')}")
                    
                    # Mark as extracted regardless of success or failure
                    attachment.last_extracted_at = datetime.now()

                except Exception as e:
                    print(f"❌ Error processing attachment {attachment.filename}: {str(e)}")
                    # Mark as extracted even if there was an error
                    attachment.last_extracted_at = datetime.now()

        db.commit() # Commit all last_extracted_at updates
        print("💾 Updated last_extracted_at timestamps in database.")

        # Generate Excel report with cumulative writing
        if all_thread_data:
            # Get current week number
            current_date = datetime.now()
            week_number = current_date.isocalendar()[1]
            
            # Get the current week's file
            current_week_file = output_dir / f"current_week_extracts.xlsx"
            
            # Check if we need to archive the existing file for a new week
            need_new_file = False
            if current_week_file.exists():
                try:
                    wb = load_workbook(current_week_file)
                    ws = wb.active
                    
                    # Check the week number from the worksheet title
                    existing_week = None
                    if ws.title and "Week" in ws.title:
                        try:
                            existing_week = int(ws.title.split("Week")[1].strip())
                        except (ValueError, IndexError):
                            pass
                    
                    # If week numbers don't match, archive the old file
                    if existing_week and existing_week != week_number:
                        year = current_date.year
                        archived_file = output_dir / f"extracts_week{existing_week}_{year}.xlsx"
                        current_week_file.rename(archived_file)
                        print(f"📦 Archived previous week's file: {archived_file}")
                        need_new_file = True
                    
                except Exception as e:
                    print(f"⚠️ Could not check existing file week, archiving as backup: {e}")
                    backup_file = output_dir / f"backup_{current_date.strftime('%Y%m%d_%H%M%S')}_extracts.xlsx"
                    current_week_file.rename(backup_file)
                    need_new_file = True
            else:
                need_new_file = True
            
            # Load existing workbook or create new one
            if current_week_file.exists() and not need_new_file:
                try:
                    wb = load_workbook(current_week_file)
                    ws = wb.active
                    
                    # Get existing headers from the first row
                    existing_headers = []
                    if ws.max_row > 0:
                        for cell in ws[1]:
                            if cell.value:
                                existing_headers.append(cell.value)
                    
                    # Get all unique entity names from current extraction + existing file
                    all_entity_names = set()
                    for thread_data in all_thread_data.values():
                        all_entity_names.update(thread_data["Entities"].keys())
                    
                    # Add existing entity columns to preserve order
                    if existing_headers:
                        for header in existing_headers[4:]:  # Skip first 4 basic columns
                            all_entity_names.add(header)
                    
                    # Create complete headers maintaining consistent order
                    headers = ["Thread ID", "Email Subject", "Email Sender", "Extraction Date"]
                    sorted_entities = sorted(all_entity_names)
                    headers.extend(sorted_entities)
                    
                    # Update headers if new columns were added
                    if len(headers) > len(existing_headers):
                        # Clear the first row and rewrite with complete headers
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=1, column=col, value=headers[col-1] if col <= len(headers) else "")
                        
                        # Update existing data rows with new columns (fill with empty strings)
                        for row_idx in range(2, ws.max_row + 1):
                            for col_idx in range(len(existing_headers) + 1, len(headers) + 1):
                                ws.cell(row=row_idx, column=col_idx, value="")
                    
                    print(f"📊 Loaded existing Excel file for week {week_number} with {ws.max_row - 1} existing records")
                    
                except Exception as e:
                    print(f"⚠️ Could not load existing file, creating new one: {e}")
                    need_new_file = True
            
            if need_new_file or not current_week_file.exists():
                # Create new workbook for the current week
                wb = Workbook()
                ws = wb.active
                ws.title = f"Week {week_number}"
                
                # Create headers for new file
                all_entity_names = set()
                for thread_data in all_thread_data.values():
                    all_entity_names.update(thread_data["Entities"].keys())
                headers = ["Thread ID", "Email Subject", "Email Sender", "Extraction Date"]
                headers.extend(sorted(all_entity_names))
                ws.append(headers)
                print(f"📄 Created new Excel file for week {week_number}")
            
            # Get existing thread data to update instead of skipping
            existing_thread_rows = {}
            current_headers = []
            if ws.max_row > 0:
                # Get current headers
                for cell in ws[1]:
                    if cell.value:
                        current_headers.append(cell.value)
                
                # Map existing thread IDs to their row numbers and data
                if ws.max_row > 1:
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if row and row[0]:  # Thread ID is in first column
                            thread_id_str = str(row[0])
                            existing_thread_rows[thread_id_str] = {
                                'row_number': row_idx,
                                'data': dict(zip(current_headers, row))
                            }
            
            # Process thread data - update existing rows or add new ones
            new_rows_added = 0
            updated_rows = 0
            
            for thread_id, thread_data in all_thread_data.items():
                thread_id_str = str(thread_id)
                
                if thread_id_str in existing_thread_rows:
                    # Update existing row with new data
                    print(f"🔄 Updating existing thread {thread_id} with new extraction data")
                    existing_row = existing_thread_rows[thread_id_str]
                    row_number = existing_row['row_number']
                    existing_data = existing_row['data']
                    
                    # Merge new entities with existing ones (new data takes precedence)
                    merged_entities = existing_data.copy()
                    
                    # Update basic fields with latest extraction
                    merged_entities["Extraction Date"] = thread_data["Extraction Date"]
                    
                    # Merge entity data - keep existing values, add new ones
                    for entity_name, entity_value in thread_data["Entities"].items():
                        if entity_value:  # Only update with non-empty values
                            merged_entities[entity_name] = entity_value
                    
                    # Update the Excel row with merged data
                    for col_idx, header in enumerate(current_headers, start=1):
                        if header == "Thread ID":
                            ws.cell(row=row_number, column=col_idx, value=thread_id)
                        elif header == "Email Subject":
                            ws.cell(row=row_number, column=col_idx, value=thread_data["Email Subject"])
                        elif header == "Email Sender":
                            ws.cell(row=row_number, column=col_idx, value=thread_data["Email Sender"])
                        else:
                            # Use merged data for all other fields
                            ws.cell(row=row_number, column=col_idx, value=merged_entities.get(header, ""))
                    
                    updated_rows += 1
                else:
                    # Add new row for new thread
                    print(f"➕ Adding new thread {thread_id} to Excel")
                    
                    row_data = []
                    for header in current_headers:
                        if header == "Thread ID":
                            row_data.append(thread_id)
                        elif header == "Email Subject":
                            row_data.append(thread_data["Email Subject"])
                        elif header == "Email Sender":
                            row_data.append(thread_data["Email Sender"])
                        elif header == "Extraction Date":
                            row_data.append(thread_data["Extraction Date"])
                        else:
                            # This is an entity column
                            row_data.append(thread_data["Entities"].get(header, ""))
                    
                    ws.append(row_data)
                    new_rows_added += 1
            
            # Save the Excel file
            wb.save(current_week_file)
            print(f"📊 Excel report updated: {current_week_file}")
            print(f"➕ Added {new_rows_added} new thread records")
            print(f"🔄 Updated {updated_rows} existing thread records with new data")
            print(f"📈 Total records in file: {ws.max_row - 1}")
        else:
            print("ℹ️ No data was successfully extracted to generate Excel reports.")

    except Exception as e:
        print(f"❌ Error in process_attachments task: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()

    finally:
        db.close()
